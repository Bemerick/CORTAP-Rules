"""
Excel Workbook Generator Service
Generates FTA Audit Workbooks from database data
"""

from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session
from typing import List, Dict, Any

from app.models import SubArea, Section, IndicatorOfCompliance, ProjectApplicability, Deficiency


def get_project_workbook_data(db: Session, project_id: int) -> Dict[str, Any]:
    """
    Fetch applicable sub-areas and their indicators for a project from the database.
    Returns data in the format expected by the workbook generator.
    """
    # Get applicable sub-areas for the project
    sub_areas = db.query(SubArea).join(
        ProjectApplicability,
        SubArea.id == ProjectApplicability.sub_area_id
    ).filter(
        ProjectApplicability.project_id == project_id,
        ProjectApplicability.is_applicable == True
    ).order_by(SubArea.section_id, SubArea.id).all()

    if not sub_areas:
        return {"sections": []}

    # Get all sections that have applicable sub-areas
    section_ids = list(set([sa.section_id for sa in sub_areas]))
    sections = db.query(Section).filter(Section.id.in_(section_ids)).all()

    # Create a map of section_id to section
    sections_map = {s.id: s for s in sections}

    # Get all indicators for these sub-areas
    sub_area_ids = [sa.id for sa in sub_areas]
    indicators = db.query(IndicatorOfCompliance).filter(
        IndicatorOfCompliance.sub_area_id.in_(sub_area_ids)
    ).all()

    # Group indicators by sub_area_id
    indicators_by_sub_area = {}
    for indicator in indicators:
        if indicator.sub_area_id not in indicators_by_sub_area:
            indicators_by_sub_area[indicator.sub_area_id] = []
        indicators_by_sub_area[indicator.sub_area_id].append({
            'indicator_id': indicator.indicator_id,
            'text': indicator.text
        })

    # Get all deficiencies for these sub-areas
    deficiencies = db.query(Deficiency).filter(
        Deficiency.sub_area_id.in_(sub_area_ids)
    ).all()

    # Group deficiencies by sub_area_id
    deficiencies_by_sub_area = {}
    for deficiency in deficiencies:
        if deficiency.sub_area_id not in deficiencies_by_sub_area:
            deficiencies_by_sub_area[deficiency.sub_area_id] = []
        deficiencies_by_sub_area[deficiency.sub_area_id].append({
            'code': deficiency.code,
            'title': deficiency.title,
            'determination': deficiency.determination,
            'suggested_corrective_action': deficiency.suggested_corrective_action
        })

    # Group sub-areas by section
    sub_areas_by_section = {}
    for sa in sub_areas:
        if sa.section_id not in sub_areas_by_section:
            sub_areas_by_section[sa.section_id] = []
        sub_areas_by_section[sa.section_id].append({
            'id': sa.id,
            'question': sa.question,
            'indicators_of_compliance': indicators_by_sub_area.get(sa.id, []),
            'deficiencies': deficiencies_by_sub_area.get(sa.id, [])
        })

    # Build the sections structure
    sections_list = []
    for section_id in sorted(sub_areas_by_section.keys(),
                            key=lambda x: (sections_map[x].chapter_number or 999, x)):
        section = sections_map[section_id]
        sections_list.append({
            'section': {
                'id': section.id,
                'title': section.title,
                'chapter_number': section.chapter_number
            },
            'sub_areas': sub_areas_by_section[section_id]
        })

    return {'sections': sections_list}


def create_workbook_bytes(data: Dict[str, Any]) -> BytesIO:
    """
    Create Excel workbook with tabs for each section.
    Returns workbook as BytesIO object.
    """
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Define styles
    header_font = Font(bold=True, size=11)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_alignment = Alignment(vertical='top', wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Define header fills
    compliant_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')  # Light green
    non_compliant_fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')  # Light red
    na_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')  # Light gray
    header_fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')  # Light blue

    # Column headers
    headers = [
        'Question',
        'Indicator of Compliance',
        'Compliant',
        'Non-Compliant',
        'N/A',
        'Audit Evidence',
        'Finding of Non-Compliance Code',
        'Audit Finding Description',
        'Additional Info',
        'Required Action'
    ]

    # Column widths
    column_widths = [50, 60, 12, 15, 10, 40, 30, 40, 40, 40]

    # Process each section
    sections = data.get('sections', [])

    for section_data in sections:
        section = section_data.get('section', {})
        section_title = section.get('title', 'Unknown')
        chapter_number = section.get('chapter_number')
        sub_areas = section_data.get('sub_areas', [])

        # Create worksheet for this section (truncate title if too long)
        # Include chapter number in sheet name if available
        if chapter_number:
            sheet_title = f"{chapter_number}. {section_title}"
        else:
            sheet_title = section_title

        # Excel sheet names are limited to 31 characters
        sheet_title = sheet_title[:31] if len(sheet_title) > 31 else sheet_title
        ws = wb.create_sheet(title=sheet_title)

        # Write headers
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border

            # Apply colored fills to compliance columns
            if header == 'Compliant':
                cell.fill = compliant_fill
            elif header == 'Non-Compliant':
                cell.fill = non_compliant_fill
            elif header == 'N/A':
                cell.fill = na_fill
            else:
                cell.fill = header_fill

        # Set column widths
        for col_idx, width in enumerate(column_widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Write data
        current_row = 2

        for sub_area in sub_areas:
            sub_area_id = sub_area.get('id', '')
            question = sub_area.get('question', '')
            indicators = sub_area.get('indicators_of_compliance', [])
            deficiencies = sub_area.get('deficiencies', [])

            # Prepend sub_area id to question
            question_with_id = f"{sub_area_id}. {question}" if sub_area_id else question

            # Prepare deficiency data for merged cells
            deficiency_codes = []
            determinations = []
            corrective_actions = []

            for idx, deficiency in enumerate(deficiencies, 1):
                code = deficiency.get('code', '')
                title = deficiency.get('title', '')
                if code and title:
                    deficiency_codes.append(f"{code} - {title}")
                elif code:
                    deficiency_codes.append(code)

                determination = deficiency.get('determination', '')
                if determination:
                    determinations.append(determination)

                action = deficiency.get('suggested_corrective_action', '')
                if action:
                    corrective_actions.append(action)

            # Join multiple deficiencies with newlines and numbering
            if len(deficiency_codes) > 1:
                deficiency_code_text = '\n\n'.join([f"{i}. {code}" for i, code in enumerate(deficiency_codes, 1)])
            else:
                deficiency_code_text = deficiency_codes[0] if deficiency_codes else ''

            if len(determinations) > 1:
                determination_text = '\n\n'.join([f"{i}. {det}" for i, det in enumerate(determinations, 1)])
            else:
                determination_text = determinations[0] if determinations else ''

            if len(corrective_actions) > 1:
                corrective_action_text = '\n\n'.join([f"{i}. {action}" for i, action in enumerate(corrective_actions, 1)])
            else:
                corrective_action_text = corrective_actions[0] if corrective_actions else ''

            # Get number of indicators for merging
            num_indicators = len(indicators) if indicators else 0

            if num_indicators > 0:
                # Merge question cell across indicator rows
                start_row = current_row
                end_row = current_row + num_indicators - 1
                ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)

                # Write question in merged cell
                cell = ws.cell(row=start_row, column=1, value=question_with_id)
                cell.alignment = cell_alignment
                cell.border = border

                # Merge columns 3-10 at question level (Compliant, Non-Compliant, N/A, Audit Evidence,
                # Finding of Non-Compliance Code, Audit Finding Description, Additional Info, Required Action)
                # We need to apply borders to all cells in merged ranges for proper display
                merged_columns_data = [
                    (3, ''),  # Compliant
                    (4, ''),  # Non-Compliant
                    (5, ''),  # N/A
                    (6, ''),  # Audit Evidence
                    (7, deficiency_code_text),  # Finding of Non-Compliance Code
                    (8, determination_text),  # Audit Finding Description
                    (9, ''),  # Additional Info
                    (10, corrective_action_text)  # Required Action
                ]

                for col_num, col_value in merged_columns_data:
                    ws.merge_cells(start_row=start_row, start_column=col_num, end_row=end_row, end_column=col_num)

                    # Set value and formatting on primary cell
                    cell = ws.cell(row=start_row, column=col_num, value=col_value)
                    cell.alignment = cell_alignment
                    cell.border = border

                    # Apply borders to all cells in the merged range
                    for row_num in range(start_row, end_row + 1):
                        cell = ws.cell(row=row_num, column=col_num)
                        cell.border = border
            else:
                # No indicators, just write question normally
                cell = ws.cell(row=current_row, column=1, value=question_with_id)
                cell.alignment = cell_alignment
                cell.border = border

                # Add deficiency data to single row
                cell = ws.cell(row=current_row, column=7, value=deficiency_code_text)
                cell.alignment = cell_alignment
                cell.border = border

                cell = ws.cell(row=current_row, column=8, value=determination_text)
                cell.alignment = cell_alignment
                cell.border = border

                cell = ws.cell(row=current_row, column=10, value=corrective_action_text)
                cell.alignment = cell_alignment
                cell.border = border

                # Add borders to empty cells in question row
                for col_idx in [2, 3, 4, 5, 6, 9]:
                    cell = ws.cell(row=current_row, column=col_idx, value='')
                    cell.border = border

                current_row += 1

            # Write indicators (skip if None)
            if indicators:
                for indicator_idx, indicator in enumerate(indicators):
                    indicator_id = indicator.get('indicator_id', '')
                    indicator_text = indicator.get('text', '')

                    # Prepend indicator_id to indicator text
                    indicator_with_id = f"{indicator_id}. {indicator_text}" if indicator_id else indicator_text

                    # Indicator text
                    cell = ws.cell(row=current_row, column=2, value=indicator_with_id)
                    cell.alignment = cell_alignment
                    cell.border = border

                    # Add borders to merged cells in this row to ensure proper border display
                    # The merged cells already have their primary cell set with borders,
                    # but we need to ensure the border appears on the indicator row
                    # Note: We can't set values on merged cells, but borders should be visible
                    # from the merge operation above

                    current_row += 1

        # Freeze top row
        ws.freeze_panes = 'A2'

    # Save workbook to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output


def generate_project_workbook(db: Session, project_id: int) -> BytesIO:
    """
    Main function to generate workbook for a project.
    Returns BytesIO object containing the Excel file.
    """
    data = get_project_workbook_data(db, project_id)
    return create_workbook_bytes(data)
